import io
import datetime
from sqlalchemy.orm import Session
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.schema import (
    Student, Class, Subject, ScheduledLecture, LectureStatus,
    AttendanceSession, AttendanceRecord, AttendanceStatus, TimetableEntry, Faculty, User
)

def calculate_student_overall_dashboard(db: Session, student_id: int):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        return None

    # Fetch all subjects for student's class
    timetable_entries = db.query(TimetableEntry).filter(TimetableEntry.class_id == student.class_id).all()
    subject_ids = list(set([t.subject_id for t in timetable_entries]))

    subject_summaries = []
    total_conducted_all = 0
    total_present_all = 0

    for sub_id in subject_ids:
        subject = db.query(Subject).filter(Subject.id == sub_id).first()
        if not subject:
            continue

        # Get faculty info for this subject
        t_entry = db.query(TimetableEntry).filter(
            TimetableEntry.class_id == student.class_id,
            TimetableEntry.subject_id == sub_id
        ).first()
        faculty_name = t_entry.faculty.user.full_name if t_entry else "Faculty"

        # Conducted lectures for this subject (COMPLETED or ACTIVE or CONFIRMED with closed session)
        conducted_lectures = db.query(ScheduledLecture).join(TimetableEntry).filter(
            TimetableEntry.class_id == student.class_id,
            TimetableEntry.subject_id == sub_id,
            ScheduledLecture.status.in_([LectureStatus.COMPLETED.value, LectureStatus.ACTIVE.value, LectureStatus.CONFIRMED.value])
        ).all()

        total_conducted = len(conducted_lectures)
        present_count = 0

        if total_conducted > 0:
            lecture_ids = [l.id for l in conducted_lectures]
            sessions = db.query(AttendanceSession).filter(AttendanceSession.scheduled_lecture_id.in_(lecture_ids)).all()
            session_ids = [s.id for s in sessions]

            if session_ids:
                present_records = db.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id.in_(session_ids),
                    AttendanceRecord.student_id == student_id,
                    AttendanceRecord.status == AttendanceStatus.PRESENT.value
                ).count()
                present_count = present_records

        pct = round((present_count / total_conducted * 100), 2) if total_conducted > 0 else 100.0

        total_conducted_all += total_conducted
        total_present_all += present_count

        subject_summaries.append({
            "subject_id": subject.id,
            "subject_code": subject.code,
            "subject_name": subject.name,
            "faculty_name": faculty_name,
            "present_count": present_count,
            "total_conducted": total_conducted,
            "percentage": pct
        })

    overall_pct = round((total_present_all / total_conducted_all * 100), 2) if total_conducted_all > 0 else 100.0

    return {
        "student": {
            "id": student.id,
            "user_id": student.user_id,
            "enrollment_no": student.enrollment_no,
            "roll_no": student.roll_no,
            "semester": student.semester,
            "branch": student.branch,
            "class_id": student.class_id,
            "name": student.user.full_name
        },
        "overall_percentage": overall_pct,
        "total_conducted": total_conducted_all,
        "total_present": total_present_all,
        "subject_summaries": subject_summaries
    }

def get_admin_dashboard_summary(db: Session, class_id: int = 1):
    cls = db.query(Class).filter(Class.id == class_id).first()
    students = db.query(Student).filter(Student.class_id == class_id).all()
    subjects = db.query(Subject).all()

    low_attendance_students = []
    total_class_pct = 0.0

    for student in students:
        summary = calculate_student_overall_dashboard(db, student.id)
        if summary:
            total_class_pct += summary["overall_percentage"]
            if summary["overall_percentage"] < 75.0:
                low_attendance_students.append({
                    "id": student.id,
                    "roll_no": student.roll_no,
                    "enrollment_no": student.enrollment_no,
                    "name": student.user.full_name,
                    "overall_percentage": summary["overall_percentage"]
                })

    avg_class_pct = round(total_class_pct / len(students), 2) if students else 0.0

    today = datetime.date.today()
    today_lectures = db.query(ScheduledLecture).filter(ScheduledLecture.date == today).all()
    conducted_today = [l for l in today_lectures if l.status in [LectureStatus.COMPLETED.value, LectureStatus.ACTIVE.value]]

    return {
        "class_name": cls.name if cls else "B.Tech-M.Tech CSE (Cyber Security)",
        "semester": cls.semester if cls else "III",
        "total_students": len(students),
        "total_subjects": len(subjects),
        "overall_class_attendance": avg_class_pct,
        "low_attendance_students": low_attendance_students,
        "today_lectures_count": len(today_lectures),
        "conducted_lectures_count": len(conducted_today)
    }

def generate_excel_attendance_sheet(db: Session, class_id: int = 1, subject_id: int = 1) -> bytes:
    """
    Generates an Excel workbook styled identically to the reference Lesson Attendance Sheet.
    """
    cls = db.query(Class).filter(Class.id == class_id).first()
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.roll_no).all()

    # Get faculty info for this subject
    t_entry = db.query(TimetableEntry).filter(
        TimetableEntry.class_id == class_id,
        TimetableEntry.subject_id == subject_id
    ).first()
    faculty_name = t_entry.faculty.user.full_name if t_entry else "Dr. Akash Thakkar"

    # Get all conducted lectures for this class and subject
    conducted_lectures = db.query(ScheduledLecture).join(TimetableEntry).filter(
        TimetableEntry.class_id == class_id,
        TimetableEntry.subject_id == subject_id,
        ScheduledLecture.status.in_([LectureStatus.COMPLETED.value, LectureStatus.ACTIVE.value, LectureStatus.CONFIRMED.value])
    ).order_by(ScheduledLecture.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_header = Font(name="Calibri", size=9, bold=True)
    font_title = Font(name="Calibri", size=12, bold=True)
    
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_present = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    fill_absent = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red

    border_thin = Side(border_style="thin", color="000000")
    border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    # Row 1: Header metadata
    ws.merge_cells("A1:C1")
    ws["A1"] = f"FACULTY NAME: {faculty_name}"
    ws["A1"].font = font_bold

    ws["D1"] = f"LESSON ATTENDANCE SHEET, {cls.name if cls else 'B.Tech Sem-III'}"
    ws["D1"].font = font_title
    ws["D1"].alignment = align_center

    ws["H1"] = "TERM: JULY-DEC 2026"
    ws["H1"].font = font_bold

    # Row 2: Metadata block
    ws["A2"] = "SUBJECT CODE"
    ws["B2"] = subject.code if subject else "CTBT-BSC-301"
    ws["C2"] = "SEM-III"
    ws["D2"] = "TERM START"
    ws["E2"] = cls.term_start.strftime("%d/%m/%Y") if cls else "22/07/2026"
    ws["F2"] = "TERM END"
    ws["G2"] = cls.term_end.strftime("%d/%m/%Y") if cls else "31/12/2026"
    ws["H2"] = "SUBJECT NAME"
    ws["I2"] = subject.name if subject else "Engineering Mathematics-III"

    for r in range(1, 3):
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.font = font_bold

    # Row 4: Grid headers
    ws["A4"] = "Sr. No."
    ws["B4"] = "ENROLLMENT NO."
    ws["C4"] = "NAME OF STUDENTS"

    ws.cell(row=4, column=1).alignment = align_center
    ws.cell(row=4, column=2).alignment = align_center
    ws.cell(row=4, column=3).alignment = align_left

    col_idx = 4
    lecture_map = {}  # col_idx -> lecture_id
    for idx, lec in enumerate(conducted_lectures, start=1):
        cell_date = ws.cell(row=3, column=col_idx)
        cell_date.value = lec.date.strftime("%d-%m-%Y")
        cell_date.font = font_header
        cell_date.alignment = align_center

        cell_lec = ws.cell(row=4, column=col_idx)
        cell_lec.value = f"Lec {idx}"
        cell_lec.font = font_header
        cell_lec.alignment = align_center
        cell_lec.fill = fill_header

        lecture_map[col_idx] = lec.id
        col_idx += 1

    # Final summary headers
    ws.cell(row=4, column=col_idx, value="Total Present").font = font_header
    ws.cell(row=4, column=col_idx).alignment = align_center
    ws.cell(row=4, column=col_idx).fill = fill_header

    ws.cell(row=4, column=col_idx + 1, value="Attendance %").font = font_header
    ws.cell(row=4, column=col_idx + 1).alignment = align_center
    ws.cell(row=4, column=col_idx + 1).fill = fill_header

    # Populate Student Rows
    current_row = 5
    for idx, student in enumerate(students, start=1):
        ws.cell(row=current_row, column=1, value=idx).alignment = align_center
        ws.cell(row=current_row, column=2, value=student.enrollment_no).alignment = align_center
        ws.cell(row=current_row, column=3, value=student.user.full_name).alignment = align_left

        present_count = 0
        total_conducted = len(conducted_lectures)

        for c_idx, lec_id in lecture_map.items():
            session = db.query(AttendanceSession).filter(AttendanceSession.scheduled_lecture_id == lec_id).first()
            status_val = "A"
            if session:
                rec = db.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id == session.id,
                    AttendanceRecord.student_id == student.id
                ).first()
                if rec and rec.status == AttendanceStatus.PRESENT.value:
                    status_val = "P"
                    present_count += 1

            cell = ws.cell(row=current_row, column=c_idx, value=status_val)
            cell.alignment = align_center
            cell.font = font_header
            if status_val == "P":
                cell.fill = fill_present
            else:
                cell.fill = fill_absent

        pct = round((present_count / total_conducted * 100), 2) if total_conducted > 0 else 100.0

        # Total Present & Pct
        cell_tp = ws.cell(row=current_row, column=col_idx, value=present_count)
        cell_tp.alignment = align_center
        cell_tp.font = font_bold

        cell_pct = ws.cell(row=current_row, column=col_idx + 1, value=f"{pct}%")
        cell_pct.alignment = align_center
        cell_pct.font = font_bold

        current_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_csv_attendance_sheet(db: Session, class_id: int = 1, subject_id: int = 1) -> str:
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.roll_no).all()
    conducted_lectures = db.query(ScheduledLecture).join(TimetableEntry).filter(
        TimetableEntry.class_id == class_id,
        TimetableEntry.subject_id == subject_id,
        ScheduledLecture.status.in_([LectureStatus.COMPLETED.value, LectureStatus.ACTIVE.value, LectureStatus.CONFIRMED.value])
    ).order_by(ScheduledLecture.date).all()

    headers = ["Sr No", "Enrollment No", "Student Name"]
    for idx, lec in enumerate(conducted_lectures, start=1):
        headers.append(f"Lec {idx} ({lec.date.strftime('%d/%m')})")
    headers.extend(["Total Present", "Total Conducted", "Attendance %"])

    rows = []
    for idx, student in enumerate(students, start=1):
        row = [idx, student.enrollment_no, student.user.full_name]
        present_count = 0
        total_conducted = len(conducted_lectures)

        for lec in conducted_lectures:
            session = db.query(AttendanceSession).filter(AttendanceSession.scheduled_lecture_id == lec.id).first()
            if session:
                rec = db.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id == session.id,
                    AttendanceRecord.student_id == student.id
                ).first()
                if rec and rec.status == AttendanceStatus.PRESENT.value:
                    row.append("P")
                    present_count += 1
                else:
                    row.append("A")
            else:
                row.append("A")

        pct = round((present_count / total_conducted * 100), 2) if total_conducted > 0 else 100.0
        row.extend([present_count, total_conducted, f"{pct}%"])
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)
    return df.to_csv(index=False)
